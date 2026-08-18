"""Silence is a fact about the analyzer, not a defect in the audited code.

Before this suite existed the meter reported its own blind spots through the grade:
`report.py` turned one undecided state into status `might`, and `might` into D. A
repository with five hundred pieces of state, four hundred and ninety-nine of them
provably finite, graded D because the analyzer could not read one call. Nothing about
the code caused that letter; our reading did.

The four invariants below are the design in
research/candidates/silence-index-for-finite-testability.md:

  1. Silence is its own measure, with the file and line of every silent site.
  2. The grade is decided on observed state only.
  3. Above a silence floor no grade is issued at all, so hiding everything from the
     analyzer produces no grade rather than a good one. Without this the second rule
     would reward obscurity, which is the whole poka-yoke.
  4. "I could not see into this library" and "I have not taught the analyzer this
     builtin" stay distinct facts: the first is a boundary the adopter can make
     readable, the second is our backlog.

Pure assertions over fixtures written to tmp_path, no mocks (Honest Code Rule 10).
"""

from __future__ import annotations

from l1_analyzer import card, report, state_bounds, state_partition

# One class, one piece of state, and one call the analyzer cannot follow. The state is
# never compared to anything unbounded and never used as an unbounded key, so nothing
# here is evidence that the code is untestable - only that the meter stopped reading.
ONE_SILENT_STATE = (
    "import openpyxl\n"
    "\n"
    "class Exporter:\n"
    "    def __init__(self):\n"
    "        self.workbook = openpyxl.Workbook()\n"
    "    def save(self, path):\n"
    "        return openpyxl.persist(self.workbook, path)\n"
)

# The same silent state, beside decided state that outnumbers it, so the silence
# fraction lands under any sane floor.
MOSTLY_DECIDED = ONE_SILENT_STATE + (
    "\n"
    "class Router:\n"
    "    def __init__(self):\n"
    "        self.enabled = False\n"
    "        self.mode = 'fast'\n"
    "        self.retries = 0\n"
    "        self.verbose = True\n"
    "        self.strict = False\n"
    "        self.debug = False\n"
    "        self.trace = False\n"
    "        self.quiet = True\n"
    "        self.dry_run = False\n"
    "    def route(self):\n"
    "        if self.enabled and self.verbose and self.strict and self.debug:\n"
    "            return 1\n"
    "        if self.trace and self.quiet and self.dry_run:\n"
    "            return 2\n"
    "        if self.mode == 'fast' and self.retries > 3:\n"
    "            return 3\n"
    "        return 0\n"
)

_HEALTHY = {k: {"band": "Healthy"} for k in ("L1.17", "L1.15", "L1.10", "L1.11", "L1.9", "L1.16")}


def _panel(l18b: dict) -> dict:
    return {"L1.18": {"band": "Healthy"}, "L1.18b": l18b, **_HEALTHY}


def _classify(tmp_path, src: str) -> dict:
    (tmp_path / "app.py").write_text(src)
    return state_bounds.classify(tmp_path, "python")


def test_silence_is_reported_with_a_file_and_line_for_every_silent_site(tmp_path):
    """Proposal part one. A share alone tells a reader nothing they can act on; the
    sites do. Every silent site is listed, not a capped sample, because the reader's
    next move is to open each one and decide whether to make it readable."""
    result = _classify(tmp_path, ONE_SILENT_STATE)
    silence = result["silence"]
    assert silence["count"] == 1
    assert 0.0 < silence["fraction"] <= 1.0
    assert len(silence["sites"]) == silence["count"]
    site = silence["sites"][0]
    assert site["file"] == "app.py"
    # Line 7, `openpyxl.persist(self.workbook, path)`, is the reference nobody could
    # read. This asserted 5, the line the state is BOUND on, until 2026-08-18. Line 5 is
    # `self.workbook = openpyxl.Workbook()`, which is not the thing that went unread and
    # tells a reader nothing to act on. The site exists so the reader can open it and
    # decide whether to model that callee, so it follows the callee.
    assert site["line"] == 7
    assert site["state"] == "self.workbook"
    assert site["reason"]
    # And the reader gets them, not just the JSON: a share nobody can act on is not a
    # measure, it is a number.
    markdown = card.card_markdown(card.build_card("x", "python", _panel(result)))
    assert "What we could not follow" in markdown
    # The reason, not only the site. Which of the four it is decides whose move is next, and
    # the card printed the location without it until report.py's renderers were deleted and
    # this assertion moved to the output that ships.
    # 5 until 2026-08-18, when the site started following the unread reference
    # rather than the binding. The rendered line is what an adopter opens.
    assert "`app.py:7` — `self.workbook` (handed to" in markdown


def test_one_silent_state_no_longer_caps_the_grade_at_d(tmp_path):
    """Proposal part two, and the whole argument. The counts that set the status are
    counts over DECIDED state. An undecided state is not evidence about the code, so it
    cannot push a repository into a lower band on its own."""
    result = _classify(tmp_path, MOSTLY_DECIDED)
    assert result["counts"]["promiscuous"] == 0
    assert result["counts"]["unresolved"] >= 1
    g = report.grade_summary(_panel(result), report.UNORDERED_CLASS_BOUND)
    assert g["status"] == "can"
    assert g["grade"] == "A"


def test_silence_above_the_floor_issues_no_grade_rather_than_a_good_one(tmp_path):
    """Proposal part three, the poka-yoke. Grading on observed state alone would hand an
    A to code that shows the analyzer nothing, because zero observed state is vacuously
    clean. Above the floor the status is `na` and no grade exists, so obscurity buys
    silence instead of a letter."""
    result = _classify(tmp_path, ONE_SILENT_STATE)
    assert result["silence"]["fraction"] > report.SILENCE_FLOOR
    g = report.grade_summary(_panel(result), report.UNORDERED_CLASS_BOUND)
    assert g["status"] == "na"
    assert g["grade"] is None


def test_unreadable_library_and_unmodeled_builtin_are_different_facts(tmp_path):
    """Proposal part four. `sorted(...)` is a name we could teach the analyzer and have
    not; a method on a third-party workbook is a boundary we genuinely cannot read. One
    is our backlog and one is a contract the adopter can make explicit. Reporting them
    as a single number tells the adopter to fix something that is ours."""
    (tmp_path / "app.py").write_text(
        "import openpyxl\n"
        "\n"
        "rows = []\n"
        "book = openpyxl.Workbook()\n"
        "\n"
        "def ordered():\n"
        "    return sorted(rows)\n"
        "\n"
        "def store():\n"
        "    return openpyxl.persist(book, 'out.xlsx')\n"
    )
    result = state_bounds.classify(tmp_path, "python")
    reasons = {s["state"]: s["reason"] for s in result["silence"]["sites"]}
    assert reasons["rows"] == state_partition.UNMODELED_CALLEE
    assert reasons["book"] == state_partition.EXTERNAL_BOUNDARY
    assert result["silence"]["by_reason"][state_partition.UNMODELED_CALLEE] == 1
    assert result["silence"]["by_reason"][state_partition.EXTERNAL_BOUNDARY] == 1


# --- a silence site must point at the shape it names -------------------------

def test_a_silence_site_points_at_the_line_whose_shape_it_names(tmp_path):
    """The site carried the state's BINDING line while `construct` came from the first
    silent reference, which is a different node somewhere else in the file. So the report
    said "at m.py:3 the shape is an attribute in a keyword argument" and line 3 is an
    assignment. _verdict's own docstring says the construct travels with its reference
    "so it names the shape at the site the reader is being sent to, and picking it from a
    different reference would send them to one place and describe another" -- which is
    what the binding line did to it.

    This matters beyond tidiness. `construct` exists so the backlog of missing rules can
    be read off the report and worked down, and a backlog whose line numbers point at the
    wrong shape cannot be worked down. It was found while doing exactly that.
    """
    # The fixture was `f(config=self.opts)` until the keyword-argument row landed and made
    # that shape readable. An f-string interpolation is still unmodelled, which is what
    # this test needs: any construct with no rule will do, and the point is the LINE.
    src = ("class Q:\n"
           "    def __init__(self):\n"
           "        self._a = {}\n"        # line 3: where the state is bound
           "    def put(self, k, v):\n"
           "        self._a[k] = v\n"
           "\n\n\n"
           "    def go(self):\n"
           "        return f'{self._a}'\n")   # line 10: the shape with no rule
    (tmp_path / "m.py").write_text(src)
    r = state_bounds.classify(tmp_path, "python")
    finding = next(f for f in r["findings"] if f["state"] == "self._a")
    site = next(s for s in r["silence"]["sites"] if s["state"] == "self._a")

    assert site["construct"] == "attribute in interpolation"
    assert site["line"] == 10, "the site must point at the reference whose shape it names"
    assert finding["line"] == 3, "the finding still reports where the state is bound"
